"""Importa o calendario de assembleias para um JSON versionado pelo sistema."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


MONTHS = [
    "Janeiro", "Fevereiro", "Mar\u00e7o", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]
EVENTS = [
    ("vencimento_parcela", "Vencimento da parcela"),
    ("adesao", "Ades\u00e3o"),
    ("vencimento_boleto_adesao", "Vencimento do boleto de ades\u00e3o"),
    ("oferta", "Oferta"),
    ("assembleia", "Assembleia"),
    ("pagamento_lance", "Pagamento do lance"),
    ("segunda_chamada", "Segunda chamada"),
]

EVENT_GUIDANCE = {
    "vencimento_parcela": "Data prevista para o vencimento da parcela do grupo.",
    "adesao": "Data prevista para formalizar a adesão ao grupo.",
    "vencimento_boleto_adesao": "Data de vencimento do boleto de adesão.",
    "oferta": "Prazo operacional para registrar a oferta de lance antes da assembleia.",
    "assembleia": "Data prevista para a assembleia mensal do grupo.",
    "pagamento_lance": "Prazo previsto para pagar o lance após a contemplação.",
    "segunda_chamada": "Data alternativa de assembleia, quando prevista pela administradora.",
    "canal_segunda_chamada": "Canal informado pela administradora para a segunda chamada.",
}

ADMINISTRATOR_GUIDANCE = {
    "ITAÚ": [
        "Não há assembleia às segundas e quintas-feiras.",
        "As datas de adesão são 5, 10, 15 e 20; se caírem em feriado ou fim de semana, passam para o próximo dia útil.",
        "A adesão ocorre dois dias antes do vencimento do boleto de adesão, desconsiderando fins de semana e feriados.",
        "A assembleia ocorre cinco dias após o vencimento do boleto de adesão, desconsiderando fins de semana e feriados.",
        "A oferta deve ser registrada até um dia antes da assembleia, desconsiderando fins de semana e feriados.",
        "O vencimento da parcela ocorre no mesmo dia do vencimento do boleto de adesão, desconsiderando fins de semana e feriados.",
        "Não há segunda chamada; o calendário apresenta um traço nesse campo.",
    ],
    "PAN": [
        "1º bloco: grupos 8019, 8020, 8021 e 8023.",
        "2º bloco: grupos 8022 e 8025.",
    ],
    "CAOA": [
        "Vencimento no dia 10: grupos 545 e 558.",
        "Vencimento no dia 20: grupos 585 e 595.",
    ],
    "EMBRACON": [
        "Vencimento no dia 10: grupos 340, 763, 342, 768, 769, 774, 788, 792, 794, 797, 742, 744 e 746.",
        "Vencimento no dia 20: grupos 761, 762, 764, 765, 766, 767, 770, 771, 772, 776, 779, 781, 785, 786, 787, 789, 790, 791, 793, 795, 727, 743, 751 e 7017.",
        "Vencimento no dia 26: grupos 338, 339, 341, 760, 775, 777, 778, 780, 782, 783, 784, 796, 749 e 7018.",
    ],
}

OFFER_GUIDANCE = [
    "O prazo varia conforme a administradora e a faixa do grupo.",
    "As referências operacionais cadastradas incluem: quinta-feira anterior à assembleia, quatro dias úteis antes, dois dias antes e um dia antes.",
    "Em grupos com vencimento de parcela no dia 10, pode haver prazo de até um dia útil após o vencimento.",
]

TEXT_REPLACEMENTS = {
    "Calend\ufffdrio": "Calend\u00e1rio",
    "IM\ufffdVEL": "IM\u00d3VEL",
    "Mar\ufffdo": "Mar\u00e7o",
    "Ades\ufffdo": "Ades\u00e3o",
    "ades\ufffdo": "ades\u00e3o",
    "assembl\ufffdia": "assembleia",
    "Assembl\ufffdia": "Assembleia",
    "Ita\ufffd": "Ita\u00fa",
    "ITA\ufffd": "ITA\u00da",
    "n\ufffdo": "n\u00e3o",
    "N\ufffdo": "N\u00e3o",
    "s\ufffdo": "s\u00e3o",
    "S\ufffdo": "S\u00e3o",
    "ap\ufffds": "ap\u00f3s",
    "Ap\ufffds": "Ap\u00f3s",
    "at\ufffd": "at\u00e9",
    "At\ufffd": "At\u00e9",
    "pr\ufffdximo": "pr\u00f3ximo",
    "Pr\ufffdximo": "Pr\u00f3ximo",
    "\ufffdteis": "\u00fateis",
    "\ufffdtil": "\u00fatil",
    "1\ufffd": "1\u00ba",
    "2\ufffd": "2\u00aa",
}


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    for source, target in TEXT_REPLACEMENTS.items():
        text = text.replace(source, target)
    return " ".join(text.split())


def normalized_administrator(value: str) -> str:
    return clean_text(value).upper().replace("ITAU", "ITAÚ")


def serialize_calendar_value(value: object, month_index: int) -> dict:
    if value in (None, ""):
        return {"display": "-", "value": None, "iso_date": None}
    if isinstance(value, (date, datetime)):
        normalized_year = 2027 if month_index == 12 and value.month == 1 else value.year
        normalized = date(normalized_year, value.month, value.day)
        return {
            "display": normalized.strftime("%d/%m/%Y"),
            "value": normalized.day,
            "iso_date": normalized.isoformat(),
        }
    if isinstance(value, (int, float)) and float(value).is_integer():
        return {"display": str(int(value)), "value": int(value), "iso_date": None}
    return {"display": clean_text(value), "value": clean_text(value), "iso_date": None}


def import_workbook(source: Path) -> dict:
    workbook = openpyxl.load_workbook(source, data_only=True)
    calendar_sheet = workbook.worksheets[0]
    rules_sheet = workbook["REGRAS"]

    schedules = []
    current_administrator = ""
    for row_number in range(4, 15):
        administrator = clean_text(calendar_sheet.cell(row_number, 1).value)
        if administrator:
            current_administrator = administrator
        faixa = calendar_sheet.cell(row_number, 2).value
        months = []
        for month_index, month_name in enumerate(MONTHS, start=1):
            start_column = 3 + ((month_index - 1) * len(EVENTS))
            events = []
            for offset, (event_id, event_label) in enumerate(EVENTS):
                cell = calendar_sheet.cell(row_number, start_column + offset)
                events.append({
                    "id": event_id,
                    "label": event_label,
                    "source_cell": cell.coordinate,
                    **serialize_calendar_value(cell.value, month_index),
                })
            months.append({"number": month_index, "name": month_name, "events": events})
        schedules.append({
            "source_row": row_number,
            "administrator": current_administrator,
            "faixa": int(faixa) if isinstance(faixa, (int, float)) else clean_text(faixa),
            "guidance": ADMINISTRATOR_GUIDANCE.get(normalized_administrator(current_administrator), []),
            "months": months,
        })

    rules = []
    current_administrator = ""
    rule_fields = [
        (3, "vencimento", "Vencimento"),
        (4, "adesao", "Ades\u00e3o"),
        (5, "oferta", "Oferta"),
        (6, "pagamento_lance", "Pagamento do lance"),
        (7, "assembleia", "Assembleia"),
        (8, "segunda_chamada", "Segunda chamada"),
        (9, "canal_segunda_chamada", "Canal da segunda chamada"),
    ]
    for row_number in range(4, 21):
        administrator = clean_text(rules_sheet.cell(row_number, 1).value)
        if administrator:
            current_administrator = administrator
        faixa = rules_sheet.cell(row_number, 2).value
        parameters = [
            {
                "id": field_id,
                "label": label,
                "value": clean_text(rules_sheet.cell(row_number, column).value) or "-",
                "source_cell": f"{get_column_letter(column)}{row_number}",
            }
            for column, field_id, label in rule_fields
        ]
        rules.append({
            "source_row": row_number,
            "administrator": current_administrator,
            "faixa": int(faixa) if isinstance(faixa, (int, float)) else clean_text(faixa),
            "guidance": ADMINISTRATOR_GUIDANCE.get(normalized_administrator(current_administrator), []),
            "parameters": parameters,
        })

    observations = []
    seen = set()
    for row_number in range(16, calendar_sheet.max_row + 1):
        for cell in calendar_sheet[row_number]:
            value = clean_text(cell.value)
            if not value or value in seen:
                continue
            seen.add(value)
            observations.append({"source_cell": cell.coordinate, "text": value})

    return {
        "metadata": {
            "title": "Calend\u00e1rio de Assembleia 2026",
            "source_file": source.name,
            "reference_year": 2026,
            "next_cycle_year": 2027,
            "date_rule": "Datas posteriores a dezembro de 2026 pertencem a janeiro de 2027.",
            "calendar_sheet": clean_text(calendar_sheet.title),
            "rules_sheet": clean_text(rules_sheet.title),
        },
        "event_types": [
            {
                "id": event_id,
                "label": label,
                "guidance": EVENT_GUIDANCE[event_id],
                "details": OFFER_GUIDANCE if event_id == "oferta" else [],
            }
            for event_id, label in EVENTS
        ],
        "guidance": {
            "administrators": {
                administrator: {"items": items}
                for administrator, items in ADMINISTRATOR_GUIDANCE.items()
            },
        },
        "schedules": schedules,
        "rules": rules,
        "observations": observations,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = import_workbook(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
